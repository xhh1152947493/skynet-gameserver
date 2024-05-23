from openpyxl import load_workbook

import os
import ast

client_json = {}
server_json = {}

pwd_current = os.getcwd()
pwd_root = os.path.abspath(os.path.join(pwd_current, "../../"))
pwd_game_root = os.path.abspath(os.path.join(pwd_root, "./Servers/game/"))
pwd_excel = os.path.abspath(os.path.join(pwd_root, "./Excels/"))
pwd_output_c = os.path.abspath(os.path.join(pwd_root, "./Code/resources/client_json/"))
pwd_output_s = os.path.abspath(os.path.join(pwd_game_root, "./config/server_json/"))
pwd_output_s_jsonmgr = os.path.abspath(os.path.join(pwd_game_root, "./config/lua/"))

def refixname(name):
	name = name.replace("_c", "")
	name = name.replace("_s", "")
	return name


def is_integer_or_float(s):
	try:
		float_value = float(s)
		return float_value.is_integer()
	except ValueError:
		return False


def process_excel_file(file_path):
	# 打开 Excel 文件
	workbook = load_workbook(file_path)

	# 遍历所有页签
	for sheet_name in workbook.sheetnames:
		sheet = workbook[sheet_name]
		fixname = refixname(sheet_name)
		if sheet_name.endswith("_s"):
			client_json[fixname] = []
		elif sheet_name.endswith("_c"):
			server_json[fixname] = []
		else:
			client_json[fixname] = []
			server_json[fixname] = []

		print(f"Dump Json From...{file_path}")

		j = -1
		# 处理每个页签的数据
		names, types = [], []
		tmp_s, tmp_c = set(), set()
		for row in sheet.iter_rows(min_row=1, values_only=True):
			j += 1
			if j == 0:
				assert row[0] != "#" "#文件头配置错误"
				continue
			if j == 1:
				assert row[0] != "$" "$文件头配置错误"
				names = row
				for i in range(1, len(names)):
					tmp_s.add(i)
					tmp_c.add(i)
					if str(names[i]).endswith("_s"):
						tmp_c.remove(i)
					if str(names[i]).endswith("_c"):
						tmp_s.remove(i)
				continue
			if j == 2:
				assert row[0] != "!" "!文件头配置错误"
				types = row
				continue
			ds, dc = {}, {}
			for i in range(1, len(names)):
				if row[i] is None:
					continue
				k = names[i]
				v = None
				t = types[i]
				if t == "string":
					v = str(row[i])
				elif t == "int":
					v = int(row[i])
				elif t == "list":
					v = ast.literal_eval(row[i])
				elif t == "float":
					v = float(row[i])
				else:
					assert f"暂不支持此类型:{t}"
				if i in tmp_s:
					ds[refixname(k)] = v
				if i in tmp_c:
					dc[refixname(k)] = v
			tmps = server_json.get(fixname, None)
			if tmps is not None:
				tmps.append(ds)
			tmpc = client_json.get(fixname, None)
			if tmpc is not None:
				tmpc.append(dc)

	# 关闭 Excel 文件
	workbook.close()


def process_excel_files_in_directory():
	print("process_excel_files_in_directory---", pwd_excel)
	# 遍历目录下的所有文件
	for filename in os.listdir(pwd_excel):
		if filename.endswith(".xlsx"):
			file_path = os.path.join(pwd_excel, filename)
			process_excel_file(file_path)


def write_json():
	import json

	# 在客户端目录下写入
	for k, v in client_json.items():
		file_path = os.path.join(pwd_output_c, f"Cfg{k}.json")
		with open(file_path, 'w', encoding="utf-8") as json_file:
			json.dump(v, json_file, ensure_ascii=False, separators=(',', ':'), indent=None)
			print(f"Write Client Json...{file_path}")
			
	# 在服务器目录下写入
	for k, v in server_json.items():
		file_path = os.path.join(pwd_output_s, f"Cfg{k}.json")
		with open(file_path, 'w', encoding="utf-8") as json_file:
			json.dump(v, json_file, ensure_ascii=False, separators=(',', ':'), indent=None)
			print(f"Write Server Json...{file_path}")

def main():
	process_excel_files_in_directory()

	write_json()

	gen_server_jsoncfg_mgr_file()

content = """
-- Generate by tools, Do not Edit.

local cfgmgr = {
%s
}

return cfgmgr
"""


def gen_server_jsoncfg_mgr_file():
	cfg = ""
	for k, v in server_json.items():
		cfg += f"\tCfg{k} = {{items = nil, onloadpost = nil}},\n"

	print(content % cfg)

	file_path = os.path.join(pwd_output_s_jsonmgr, f"jsoncfg_def.lua")
	with open(file_path, "w", encoding="utf-8") as f:
		f.write(content % cfg)
		print(f"Gen Server JsonCfg...{file_path}")


if __name__ == "__main__":
	main()
